import os
import time
from datetime import date, datetime
from json import load, dump
from json.decoder import JSONDecodeError
from typing import List

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def log_data(data: List[dict], gender: str, query: str):
    # excel log
    ROOT_DIR = os.getcwd()
    EXCEL_PATH = f"{ROOT_DIR}/logs/American Eagle/spreadsheet/items_{gender}.xlsx"
    JSON_PATH = f"{ROOT_DIR}/logs/American Eagle/json/items_{gender}.json"
    STARTING_ROW = 6
    TODAYS_DATE = date.today()

    if not os.path.isfile(EXCEL_PATH):
        wb_new = Workbook()
        sheet = wb_new.active
        customize_sheet(sheet, query)
        wb_new.save(EXCEL_PATH)

    # get all items in the item column
    wb_existing = load_workbook(filename=EXCEL_PATH)
    # assign current sheet to log
    curr_sheet = None
    for sheet in wb_existing.worksheets:
        if sheet.title == query:
            curr_sheet = sheet
    if curr_sheet is None:
        curr_sheet = wb_existing.create_sheet()
        customize_sheet(curr_sheet, query)

    old_items = {}
    # try to access json log to retrieve positions of already logged items in the spreadsheet
    try:
        with open(JSON_PATH) as f:
            items = load(f)
            for item in items:
                old_items[item["item_name"]] = {
                    "sheet": item["log_pos"]["sheet"],
                    "row": item["log_pos"]["row"]
                }
    except FileNotFoundError:
        pass

    same_day = TODAYS_DATE == datetime.strptime(curr_sheet["B1"].value, "%Y-%m-%d").date()
    next_column = curr_sheet.max_column if same_day else curr_sheet.max_column + 1
    # add the days prices to sheet
    for item in data:
        curr_item_name = item["item_name"]
        write_row = curr_sheet.max_row + 1
        if curr_item_name in old_items:
            if curr_sheet.title == old_items[curr_item_name]["sheet"]:
                write_row = old_items[curr_item_name]["row"]
        else:
            item["log_pos"] = {
                "sheet": curr_sheet.title,
                "row": write_row
            }
        curr_sheet.cell(row=write_row, column=1).value = curr_item_name
        curr_sheet.column_dimensions[
            get_column_letter(next_column)].width = 20  # if next_column is 2, B is returned from chr()
        curr_sheet.cell(row=STARTING_ROW - 1, column=next_column).value = f"Prices for {TODAYS_DATE}"

        normal_price = item.get("price").get("regular_price")
        sale_price = item.get("price").get("sale_price")
        if sale_price is None:
            curr_sheet.cell(row=write_row, column=next_column).value = normal_price
        else:
            curr_sheet.cell(row=write_row, column=next_column).value = sale_price
            curr_sheet.cell(row=write_row, column=next_column).font = Font(color="ff92d050")  # aRGB value

    # highlight the lowest value in each row
    for row in curr_sheet.iter_rows(6, curr_sheet.max_row, 2, curr_sheet.max_column):
        for cell in row:
            cell.fill = PatternFill()
        min_cell = min(reversed(row), key=lambda c: float(c.value) if c.value is not None else 9999)
        highlight_fill = PatternFill(fill_type="solid", fgColor="ff963634")
        min_cell.fill = highlight_fill
    curr_sheet["B1"].value = str(TODAYS_DATE)
    save(wb_existing, EXCEL_PATH, 5)

    with open(JSON_PATH, "r+", opener=lambda name, flags: os.open(name, flags | os.O_CREAT)) as f:
        try:
            current_data = load(f)
            current_item_names = [item.get("item_name") for item in current_data]
            for new_item in data:
                try:
                    match = current_item_names.index(new_item["item_name"])
                    # just update the prices
                    current_data[match].update(price=new_item.get("price"))
                except ValueError:
                    # tells us the item was not logged before
                    current_data.append(new_item)
            f.seek(0)  # get back to start to delete
            f.truncate(0)
            f.seek(0)  # get back to start to write
            dump(current_data, f, indent=8)
        except JSONDecodeError:
            # new file
            dump(data, f, indent=8)


def customize_sheet(sheet, title: str):
    sheet["A1"] = "Last updated:"
    sheet["B1"] = str(date.today())
    sheet["A5"] = "Item Name"
    sheet.title = title
    sheet.column_dimensions["A"].width = 50


def save(wb: Workbook, filename: str, timeout: int):
    try:
        wb.save(filename)
    except PermissionError:
        print("Please close Excel if it is currently open")
        time.sleep(timeout)
        save(wb, filename, timeout)
